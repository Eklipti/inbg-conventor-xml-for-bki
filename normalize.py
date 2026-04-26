import sys
from pathlib import Path

import openpyxl
from loguru import logger

from utils import convert_xls_to_xlsx, validate_excel


def process_excel_returns(file_path: str | Path) -> Path | None:
    """Обрабатывает Excel-файл: агрегирует возвраты и пересчитывает остаток долга.

    Ищет целевые заголовки во второй строке листа "Активные". Группирует суммы
    по ключу ("Ключевое поле", "Дата последнего возврата"). Оставляет первую
    найденную строку для каждой уникальной группы, записывает в неё общую сумму
    возврата, а значения во всех ячейках дублирующих строк очищает.

    Затем вычисляет новый остаток долга:
    - Если долг погашен в ноль: устанавливает статус "close".
    - Если долг отрицательный (перевозврат): устанавливает статус "close",
      записывает сумму в колонку "Перевозврат" и обнуляет остаток ("0,00").
    - В остальных случаях: записывает обновленный остаток долга.

    Args:
        file_path: Путь к исходному файлу Excel (строка или объект Path).

    Returns:
        Путь к обработанному файлу (Path) при успешном завершении или None,
        если файл не прошел валидацию или не найдены нужные заголовки.

    Raises:
        SystemExit: При возникновении критических ошибок в процессе обработки
            или проблем с преобразованием числовых данных.
    """
    path_obj = Path(file_path)

    try:
        if path_obj.suffix.lower() == ".xls":
            logger.info("Обнаружен формат .xls, запуск конвертации в .xlsx.")
            path_obj = convert_xls_to_xlsx(path_obj)
        else:
            logger.debug(f"Файл имеет формат {path_obj.suffix}, конвертация не требуется.")

        if not validate_excel(path_obj):
            logger.warning(f"Файл {path_obj.name} не прошел валидацию.")
            return None

        logger.debug("Загрузка рабочей книги Excel.")
        workbook = openpyxl.load_workbook(path_obj)
        sheet = workbook["Активные"]

        target_headers: list[str] = [
            "Ключевое поле",
            "Сумма последнего возрата",
            "Дата последнего возврата",
            "Остаток долга",
            "Статус долга (Операция)",
            "Перевозврат",
        ]
        header_indices: dict[str, int] = {}

        logger.trace("Поиск индексов заголовков во второй строке.")
        for row in sheet.iter_rows(min_row=2, max_row=2):
            for cell in row:
                if cell.value in target_headers:
                    header_indices[cell.value] = cell.column - 1

        if len(header_indices) != len(target_headers):
            missing = set(target_headers) - set(header_indices.keys())
            logger.error(f"Не найдены заголовки: {missing}. Найдено: {list(header_indices.keys())}")
            return None

        idx_key = header_indices["Ключевое поле"]
        idx_sum: int = header_indices["Сумма последнего возрата"]
        idx_date = header_indices["Дата последнего возврата"]
        idx_debt: int = header_indices["Остаток долга"]
        idx_status: int = header_indices["Статус долга (Операция)"]
        idx_overpayment: int = header_indices["Перевозврат"]

        aggregated_sums: dict[tuple[str, str], float] = {}
        first_seen_row_idx: dict[tuple[str, str], int] = {}

        logger.info("Начало итерации по строкам данных.")
        total_rows_processed = 0
        for row_idx, row in enumerate(sheet.iter_rows(min_row=3), start=3):
            cell_key = row[idx_key].value
            cell_sum = row[idx_sum].value
            cell_date = row[idx_date].value

            key_val = str(cell_key).strip() if cell_key is not None else ""
            date_val = str(cell_date).strip() if cell_date is not None else ""
            sum_str = str(cell_sum).strip() if cell_sum is not None else "0"

            if not key_val:
                logger.trace(f"Пропуск пустой строки на позиции {row_idx}.")
                continue

            try:
                clean_sum = sum_str.replace(" ", "").replace(",", ".")
                sum_float = float(clean_sum)
            except ValueError:
                logger.error(f"Ошибка в строке {row_idx}: невозможно преобразовать '{sum_str}' в число.")
                return None

            dict_key = (key_val, date_val)
            if dict_key in aggregated_sums:
                aggregated_sums[dict_key] += sum_float
                logger.trace(f"Дубликат в строке {row_idx} ({dict_key}): суммирование и очистка ячеек.")
                for cell in row:
                    cell.value = None
            else:
                aggregated_sums[dict_key] = sum_float
                first_seen_row_idx[dict_key] = row_idx
            total_rows_processed += 1

        logger.debug(
            f"Агрегация завершена. "
            f"Всего строк обработано: {total_rows_processed}. "
            f"Уникальных записей (осталось): {len(aggregated_sums)}."
        )

        logger.info("Начало записи агрегированных сумм и пересчета остатка долга.")
        for (key, date), total_sum in aggregated_sums.items():
            row_idx = first_seen_row_idx[(key, date)]

            formatted_sum: str = f"{total_sum:.2f}".replace(".", ",")
            sheet.cell(row=row_idx, column=idx_sum + 1, value=formatted_sum)

            raw_debt = sheet.cell(row=row_idx, column=idx_debt + 1).value
            debt_str: str = str(raw_debt).strip() if raw_debt is not None else "0"

            try:
                clean_debt: str = debt_str.replace(" ", "").replace(",", ".")
                current_debt: float = float(clean_debt)
            except ValueError:
                logger.error(f"Ошибка в строке {row_idx}: невозможно преобразовать остаток долга '{debt_str}' в число.")
                return None

            new_debt: float = current_debt - total_sum

            if new_debt == 0:
                logger.trace(f"Строка {row_idx}: Долг полностью погашен (в ноль).")
                sheet.cell(row=row_idx, column=idx_status + 1, value="close")

                formatted_new_debt = f"{new_debt:.2f}".replace(".", ",")
                sheet.cell(row=row_idx, column=idx_debt + 1, value=formatted_new_debt)

            elif new_debt < 0:
                logger.trace(f"Строка {row_idx}: Долг закрыт, есть перевозврат: {new_debt}.")
                sheet.cell(row=row_idx, column=idx_status + 1, value="close")

                formatted_overpayment: str = f"{new_debt:.2f}".replace(".", ",")
                sheet.cell(row=row_idx, column=idx_overpayment + 1, value=formatted_overpayment)
                sheet.cell(row=row_idx, column=idx_debt + 1, value="0,00")

            else:
                logger.trace(f"Строка {row_idx}: Обновлен остаток долга: {new_debt}.")
                formatted_new_debt = f"{new_debt:.2f}".replace(".", ",")
                sheet.cell(row=row_idx, column=idx_debt + 1, value=formatted_new_debt)

        logger.debug(f"Сохранение изменений в файл: {path_obj.name}")
        workbook.save(path_obj)
        logger.success("Нормализация успешно завершена.")

        return path_obj

    except Exception as e:
        logger.critical(f"Критический сбой при обработке файла: {e}")
        logger.exception("Стек вызовов:")
        sys.exit(1)
