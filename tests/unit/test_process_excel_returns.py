import shutil
from pathlib import Path

import openpyxl
import pytest

from normalize import process_excel_returns


def test_process_excel_returns(tmp_path: Path):
    """Тестирует нормализацию и суммирование возвратов в Excel-файле."""
    current_dir = Path(__file__).resolve().parent
    project_root = current_dir.parent.parent

    input_file = project_root / "tests" / "examples" / "excel" / "тестовый_Реестр(проверка_нормализации)-было.xls"
    expected_file = project_root / "tests" / "examples" / "excel" / "тестовый_Реестр(проверка_нормализации)-стало.xlsx"

    assert input_file.exists(), f"Входной файл не найден: {input_file}"
    assert expected_file.exists(), f"Ожидаемый файл не найден: {expected_file}"

    test_input_path = tmp_path / input_file.name
    shutil.copy(input_file, test_input_path)

    result_path = process_excel_returns(test_input_path)

    assert result_path is not None, "Функция вернула None (ошибка при выполнении)"
    assert result_path.exists(), "Итоговый файл не был создан/сохранен"
    assert result_path.suffix == ".xlsx", "Результирующий файл должен иметь формат .xlsx"

    wb_result = openpyxl.load_workbook(result_path, data_only=True)
    wb_expected = openpyxl.load_workbook(expected_file, data_only=True)

    sheet_result = wb_result["Активные"]
    sheet_expected = wb_expected["Активные"]

    assert sheet_result.max_row == sheet_expected.max_row, "Количество строк не совпадает"
    assert sheet_result.max_column == sheet_expected.max_column, "Количество колонок не совпадает"

    for row_res, row_exp in zip(sheet_result.iter_rows(values_only=True), sheet_expected.iter_rows(values_only=True)):
        assert row_res == row_exp, f"Несовпадение данных в строке. Получено: {row_res}, Ожидалось: {row_exp}"

    wb_result.close()
    wb_expected.close()
