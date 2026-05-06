import json
import pathlib
import re
import sys
from pathlib import Path

import openpyxl
from loguru import logger
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.utils import validate_excel


def process_excel_returns(returns_file_path: Path, main_file_path: Path) -> tuple[Path | None, int, int]:
    """Извлекает данные из файла возвратов, агрегирует их и обновляет основной файл.

    Читает файл возвратов (лист "взносы", любой регистр), ищет заголовки.
    Агрегирует суммы платежей по ключу (Номер ДО, Дата платежа).
    Затем открывает основной файл (лист "Активные"), ищет соответствующие
    записи по "Ключевое поле" (совпадает с "Номер ДО"). В найденную строку
    записывается дата, общая сумма возврата для группы и пересчитывается
    остаток долга (включая установку статуса "close" и обработку перевозврата).

    Args:
        returns_file_path (Path): Путь к файлу возвратов Excel (файл-донор).
        main_file_path (Path): Путь к основному файлу Excel (файл-реципиент).

    Returns:
        tuple[Path | None, int, int]: Кортеж, содержащий:
            - Путь к обработанному основному файлу (или None при ошибке).
            - Количество успешно обновленных записей (int).
            - Количество ненайденных записей (int).

    Raises:
        SystemExit: При возникновении критических ошибок в процессе обработки.
    """
    logger.debug(f'Слияние данных из "{returns_file_path.name}" в "{main_file_path.name}"')

    ret_path_obj = Path(returns_file_path)
    main_path_obj = Path(main_file_path)

    try:
        ret_wb = openpyxl.load_workbook(ret_path_obj, data_only=True)

        ret_sheet = None
        for sheet_name in ret_wb.sheetnames:
            if sheet_name.lower() == "взносы":
                ret_sheet = ret_wb[sheet_name]
                break

        if ret_sheet is None:
            logger.error('Лист "взносы" не найден в файле возвратов.')
            return None, 0, 0

        ret_target_headers: list[str] = [
            "Долговое обязательство.Номер ДО",
            "Поступление платежа.Сумма платежа",
            "Поступление платежа.Дата платежа",
        ]
        ret_headers_idx: dict[str, int] = {}

        for r_idx in [1]:
            for row in ret_sheet.iter_rows(min_row=r_idx, max_row=r_idx):
                for cell in row:
                    if cell.value and str(cell.value).strip() in ret_target_headers:
                        ret_headers_idx[str(cell.value).strip()] = cell.column - 1
            if len(ret_headers_idx) == len(ret_target_headers):
                break

        if len(ret_headers_idx) != len(ret_target_headers):
            missing = set(ret_target_headers) - set(ret_headers_idx.keys())
            logger.error(f"В файле возвратов не найдены заголовки: {missing}.")
            return None, 0, 0

        idx_ret_key: int = ret_headers_idx["Долговое обязательство.Номер ДО"]
        idx_ret_sum: int = ret_headers_idx["Поступление платежа.Сумма платежа"]
        idx_ret_date: int = ret_headers_idx["Поступление платежа.Дата платежа"]

        aggregated_sums: dict[tuple[str, str], float] = {}

        logger.info("Начало слияния возвратов.")
        for row_idx, row in enumerate(ret_sheet.iter_rows(min_row=3), start=3):
            cell_key = row[idx_ret_key].value
            cell_sum = row[idx_ret_sum].value
            cell_date = row[idx_ret_date].value

            key_val = str(cell_key).strip() if cell_key is not None else ""
            date_val = str(cell_date).strip() if cell_date is not None else ""
            sum_str = str(cell_sum).strip() if cell_sum is not None else "0"

            if not key_val:
                continue

            try:
                clean_sum = sum_str.replace(" ", "").replace(",", ".")
                sum_float = float(clean_sum)
            except ValueError:
                logger.error(f"Файл возвратов, строка {row_idx}: невозможно преобразовать '{sum_str}' в число.")
                return None, 0, 0

            dict_key = (key_val, date_val)
            if dict_key in aggregated_sums:
                aggregated_sums[dict_key] += sum_float
            else:
                aggregated_sums[dict_key] = sum_float

        logger.debug(f"Агрегация возвратов завершена. Уникальных записей: {len(aggregated_sums)}.")

        if not validate_excel(main_path_obj):
            logger.warning(f"Основной файл {main_path_obj.name} не прошел валидацию.")
            return None, 0, 0

        main_wb = openpyxl.load_workbook(main_path_obj)
        main_sheet = main_wb["Активные"]

        main_target_headers: list[str] = [
            "Ключевое поле",
            "Сумма последнего возрата",
            "Дата последнего возврата",
            "Остаток долга",
            "Статус долга (Операция)",
            "Перевозврат",
        ]
        main_headers_idx: dict[str, int] = {}

        for row in main_sheet.iter_rows(min_row=2, max_row=2):
            for cell in row:
                if cell.value and str(cell.value).strip() in main_target_headers:
                    main_headers_idx[str(cell.value).strip()] = cell.column - 1

        if len(main_headers_idx) != len(main_target_headers):
            missing = set(main_target_headers) - set(main_headers_idx.keys())
            logger.error(f"В основном файле не найдены заголовки: {missing}.")
            return None, 0, 0

        idx_main_key: int = main_headers_idx["Ключевое поле"]
        idx_main_sum: int = main_headers_idx["Сумма последнего возрата"]
        idx_main_date: int = main_headers_idx["Дата последнего возврата"]
        idx_main_debt: int = main_headers_idx["Остаток долга"]
        idx_main_status: int = main_headers_idx["Статус долга (Операция)"]
        idx_main_overpayment: int = main_headers_idx["Перевозврат"]

        error_sheet_name = "Не найдено в реестре"
        if error_sheet_name in main_wb.sheetnames:
            main_wb.remove(main_wb[error_sheet_name])
            logger.trace(f"Лист '{error_sheet_name}' очищен для новой итерации.")

        err_sheet = main_wb.create_sheet(error_sheet_name)
        err_sheet.append(["Источник", "Ключ", "Дата", "Сумма / Группа"])

        main_key_row_map: dict[str, int] = {}
        for row_idx, row in enumerate(main_sheet.iter_rows(min_row=3), start=3):
            cell_key = row[idx_main_key].value
            key_val = str(cell_key).strip() if cell_key is not None else ""
            if key_val and key_val not in main_key_row_map:
                main_key_row_map[key_val] = row_idx

        success_count = 0
        not_found_count = 0

        logger.info("Начало записи агрегированных сумм в основной файл.")
        for (key, date), total_sum in aggregated_sums.items():
            if key not in main_key_row_map:
                logger.warning(f"{key}: не найден.")
                err_sheet.append(["Возвраты", key, date, total_sum])
                not_found_count += 1
                continue

            success_count += 1
            row_idx = main_key_row_map[key]

            formatted_sum: str = f"{total_sum:.2f}".replace(".", ",")
            main_sheet.cell(row=row_idx, column=idx_main_sum + 1, value=formatted_sum)
            main_sheet.cell(row=row_idx, column=idx_main_date + 1, value=date)

            raw_debt = main_sheet.cell(row=row_idx, column=idx_main_debt + 1).value
            debt_str: str = str(raw_debt).strip() if raw_debt is not None else "0"

            try:
                clean_debt: str = debt_str.replace(" ", "").replace(",", ".")
                current_debt: float = float(clean_debt)
            except ValueError:
                logger.error(
                    f"Ошибка в строке {row_idx} основного файла: невозможно преобразовать остаток '{debt_str}'."
                )
                return None, 0, 0

            new_debt: float = current_debt - total_sum

            if new_debt == 0:
                logger.trace(f"{key}: Долг полностью погашен (в ноль).")
                main_sheet.cell(row=row_idx, column=idx_main_status + 1, value="close")
                formatted_new_debt = f"{new_debt:.2f}".replace(".", ",")
                main_sheet.cell(row=row_idx, column=idx_main_debt + 1, value=formatted_new_debt)

            elif new_debt < 0:
                logger.trace(f"{key}: Перевозврат - {new_debt:.2f}.")
                main_sheet.cell(row=row_idx, column=idx_main_status + 1, value="close")
                formatted_overpayment: str = f"{new_debt:.2f}".replace(".", ",")
                main_sheet.cell(row=row_idx, column=idx_main_overpayment + 1, value=formatted_overpayment)
                main_sheet.cell(row=row_idx, column=idx_main_debt + 1, value="0,00")

            else:
                logger.trace(f"{key}: Обновлен остаток долга - {current_debt:.2f} (было) / {new_debt:.2f} (стало).")
                formatted_new_debt = f"{new_debt:.2f}".replace(".", ",")
                main_sheet.cell(row=row_idx, column=idx_main_debt + 1, value=formatted_new_debt)

        logger.debug(f"Сохранение изменений в основной файл: {main_path_obj.name}")
        main_wb.save(main_path_obj)
        logger.success("Слияние возвратов успешно завершено.")

        return main_path_obj, success_count, not_found_count

    except Exception as e:
        logger.critical(f"Критический сбой при слиянии файлов: {e}")
        logger.exception("Стек вызовов:")
        sys.exit(1)


def process_other_closures(returns_file_path: Path, main_file_path: Path | None) -> tuple[Path | None, int, int]:
    """Обрабатывает иные закрытия на основе данных из специального листа.

    Ищет лист "закрытие иное" в файле возвратов. На первой строке находит
    столбцы "Объект" и "Группа ДО", собирает из них данные. Затем в основном
    файле (лист "Активные") обнуляет остаток долга для найденных объектов
    и устанавливает статус из dictionary.json согласно их группе.

    Args:
        returns_file_path (Path): Путь к файлу возвратов (источник).
        main_file_path (Path): Путь к основному файлу реестра (цель).

    Returns:
        tuple[Path | None, int, int]: Кортеж, содержащий:
            - Путь к измененному основному файлу (или None при ошибке).
            - Количество успешно обработанных закрытий (int).
            - Количество ненайденных закрытий (int).
    """
    logger.info('Запуск обработки "закрытие иное".')
    logger.debug(f"Для {returns_file_path}")

    if not main_file_path:
        logger.critical("Путь к основному файлу не передан (получен None).")
        return None, 0, 0

    try:
        project_root = Path(__file__).resolve().parent.parent
        dict_path = project_root / "dictionary.json"

        if not dict_path.exists():
            logger.critical(f"Файл словаря не найден в корне проекта: {dict_path}")
            return None, 0, 0

        with dict_path.open(encoding="utf-8") as f:
            status_mapping = json.load(f)

        ret_wb = openpyxl.load_workbook(returns_file_path, data_only=True)
        ret_sheet: Worksheet | None = None

        for sheet_name in ret_wb.sheetnames:
            if sheet_name.lower() == "закрытие иное":
                ret_sheet = ret_wb[sheet_name]
                break

        if not ret_sheet:
            logger.debug("Лист 'закрытие иное' отсутствует.")
            return None, 0, 0

        idx_ret_obj = None
        idx_ret_group = None

        first_row = next(ret_sheet.iter_rows(min_row=1, max_row=1))
        for cell in first_row:
            val = str(cell.value).strip() if cell.value else ""
            if val == "Объект":
                idx_ret_obj = cell.column
            elif val == "Группа ДО":
                idx_ret_group = cell.column

        if not idx_ret_obj or not idx_ret_group:
            logger.error("Заголовки 'Объект' или 'Группа ДО' не найдены на первой строке.")
            return None, 0, 0

        closures_to_process = {}
        for row_idx in range(2, ret_sheet.max_row + 1):
            obj_val = str(ret_sheet.cell(row_idx, idx_ret_obj).value or "").strip()
            group_val = str(ret_sheet.cell(row_idx, idx_ret_group).value or "").strip()

            if obj_val and obj_val != "None":
                closures_to_process[obj_val] = group_val

        if not closures_to_process:
            logger.warning("Данные для обработки не найдены под заголовками.")
            return None, 0, 0

        main_wb = openpyxl.load_workbook(main_file_path)

        error_sheet_name = "Не найдено в реестре"
        if error_sheet_name in main_wb.sheetnames:
            err_sheet = main_wb[error_sheet_name]
            logger.trace(f"Лист '{error_sheet_name}' найден, данные будут добавлены.")
        else:
            err_sheet = main_wb.create_sheet(error_sheet_name)
            err_sheet.append(["Источник", "Ключ", "Дата", "Сумма / Группа"])
            logger.debug(f"Лист '{error_sheet_name}' не найден, создан новый.")

        main_sheet = main_wb["Активные"]

        headers = {str(cell.value).strip(): cell.column for cell in main_sheet[2] if cell.value}
        idx_key = headers.get("Ключевое поле")
        idx_debt = headers.get("Остаток долга")
        idx_status = headers.get("Статус долга (Операция)")

        if not all([idx_key, idx_debt, idx_status]):
            logger.error("В основном файле отсутствуют необходимые столбцы.")
            return None, 0, 0

        processed_count = 0
        found_keys = set()
        for row_idx in range(3, main_sheet.max_row + 1):
            cell_val = str(main_sheet.cell(row_idx, idx_key).value or "").strip()

            if cell_val in closures_to_process:
                group_do = closures_to_process[cell_val]
                new_status = status_mapping.get(group_do, f"Неизвестный статус: {group_do}")

                main_sheet.cell(row_idx, idx_debt).value = "0,00"
                main_sheet.cell(row_idx, idx_status).value = new_status
                processed_count += 1
                found_keys.add(cell_val)
                logger.trace(f"{cell_val}: закрыт ({new_status})")

        for obj_val, group_val in closures_to_process.items():
            if obj_val not in found_keys:
                logger.warning(f"{obj_val}: не найден.")
                err_sheet.append(["Иные закрытия", obj_val, "-", group_val])

        if processed_count > 0 or len(found_keys) < len(closures_to_process):
            main_wb.save(main_file_path)
            not_found_count = len(closures_to_process) - processed_count
            logger.success("Обработка завершена.")
            return main_file_path, processed_count, not_found_count

        logger.warning("Совпадений по 'Ключевое поле' не найдено в реестре.")
        return None, 0, 0

    except Exception as e:
        logger.error(f"Критическая ошибка при обработке иных закрытий: {e}")
        logger.exception("Стек вызовов:")
        sys.exit(2)


def customize_excel(file_path: pathlib.Path | str) -> bool:
    """Кастомизирует Excel файл: применяет стили, сортировку, автофильтры и ширину колонок.

    Функция выполняет комплексную предобработку файла:
    1. Форматирует заголовки (шрифт, заливка, границы).
    2. Сортирует данные по статусу операции и наличию возвратов.
    3. Очищает и нормализует форматы дат до 'DD.MM.YYYY'.
    4. Раскрашивает строки в зависимости от логической группы (новые, закрытые, возвраты).
    5. Настраивает ширину колонок (фиксированную или по содержимому).
    6. Устанавливает автофильтры.

    Args:
        file_path (pathlib.Path | str): Путь к обрабатываемому файлу Excel.
            Должен указывать на существующий файл .xlsx.

    Returns:
        bool: True, если кастомизация прошла успешно. False, если файл не найден
            или в процессе обработки возникло критическое исключение.
    """

    try:
        path = pathlib.Path(file_path)
        if not path.is_file():
            logger.error(f"Файл не найден по указанному пути: {path}")
            return False

        wb = load_workbook(path)
        ws = wb.active

        green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        arial_font = Font(name="Arial", size=10)
        date_alignment = Alignment(horizontal="right")

        date_columns = {
            "Дата выдачи",
            "Дата рождения",
            "Дата согласия на обработку ПДН (дата договора)",
            "Дата создания (дата передачи цессии)",
            "Дата последнего возврата",
        }

        green_headers = {
            "Фамилия",
            "Имя",
            "Отчество",
            "Дата рождения",
            "Место рождения",
            "Серия-Номер",
            "Дата выдачи",
            "Кем выдан",
            "Код подразделения",
            "Уникальный идентификатор договора (сделки) БАНКА",
            "Дата согласия на обработку ПДН (дата договора)",
            "Статус долга (Операция)",
            "Дата создания (дата передачи цессии)",
            "Общая сумма долга",
            "Остаток долга",
            "Сумма последнего возрата",
            "Дата последнего возврата",
        }
        yellow_headers = {"Перевозврат", "ВОЗВРАТ"}

        headers = {}
        for cell in ws[2]:
            if cell.value:
                headers[str(cell.value)] = cell.column

            if cell.value in green_headers:
                cell.fill = green_fill
            elif cell.value in yellow_headers:
                cell.fill = yellow_fill

        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        bold_arial_font = Font(name="Arial", size=10, bold=True)
        ws.row_dimensions[2].height = 51

        headers = {}
        for cell in ws[2]:
            cell.font = bold_arial_font
            cell.border = thin_border

            if cell.value:
                headers[str(cell.value)] = cell.column

            if cell.value in green_headers:
                cell.fill = green_fill
            elif cell.value in yellow_headers:
                cell.fill = yellow_fill

        status_col = headers.get("Статус долга (Операция)")
        return_col = headers.get("Перевозврат")
        num_col = headers.get("# в33")

        rows_data = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if any(cell is not None for cell in row):
                rows_data.append(list(row))

        def _sort_key(row):
            """Определяет приоритет сортировки для строки данных.

            Args:
                row (list): Список значений ячеек текущей строки.

            Returns:
                int: Числовой код группы (0 - add, 1 - обычные, 2 - возврат, 3 - close).
            """
            status = str(row[status_col - 1]).strip().lower() if status_col and row[status_col - 1] is not None else ""
            perevozrat = row[return_col - 1] if return_col and len(row) >= return_col else 0

            try:
                p_val = float(perevozrat)
            except (ValueError, TypeError):
                p_val = 0 if perevozrat in (None, "", "0") else 1

            if status == "add":
                return 0
            if p_val != 0:
                return 2
            if status.startswith("close") or status.startswith("сlose"):
                return 3
            return 1

        sorted_rows = sorted(rows_data, key=_sort_key)

        add_fill = PatternFill(start_color="D8E4BC", end_color="D8E4BC", fill_type="solid")
        close_fill = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")
        return_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        def _parse_and_clean_date(val):
            """Принудительно извлекает дату и возвращает строку строго в формате DD.MM.YYYY.

            Удаляет временную составляющую (часы, минуты) и обрабатывает различные
            входящие форматы (ISO, RU, Slash).

            Args:
                val (Any): Исходное значение ячейки (объект datetime, строка или None).

            Returns:
                str | Any: Дата в виде строки 'DD.MM.YYYY' при успешном парсинге,
                    исходное значение в противном случае.
            """

            if val is None:
                return val

            if hasattr(val, "strftime"):
                return val.strftime("%d.%m.%Y")

            if isinstance(val, str):
                val_str = val.strip()

                match_ru = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", val_str)
                if match_ru:
                    return f"{match_ru.group(1)}.{match_ru.group(2)}.{match_ru.group(3)}"

                match_iso = re.search(r"(\d{4})-(\d{2})-(\d{2})", val_str)
                if match_iso:
                    return f"{match_iso.group(3)}.{match_iso.group(2)}.{match_iso.group(1)}"

                match_slash = re.search(r"(\d{2})/(\d{2})/(\d{4})", val_str)
                if match_slash:
                    return f"{match_slash.group(1)}.{match_slash.group(2)}.{match_slash.group(3)}"

            return val

        date_col_indices = {headers.get(col) for col in date_columns if col in headers}

        for r_idx, row_vals in enumerate(sorted_rows, start=3):
            group = _sort_key(row_vals)

            if num_col:
                row_vals[num_col - 1] = r_idx - 2

            for c_idx, val in enumerate(row_vals, start=1):
                if c_idx in date_col_indices and val is not None:
                    val = _parse_and_clean_date(val)

                cell = ws.cell(row=r_idx, column=c_idx)
                cell.value = val
                cell.border = thin_border
                cell.font = arial_font

                if c_idx in date_col_indices:
                    cell.number_format = "DD.MM.YYYY"
                    cell.alignment = date_alignment

                if group == 0:
                    cell.fill = add_fill
                elif group == 2:
                    cell.fill = return_fill
                elif group == 3:
                    cell.fill = close_fill
                else:
                    cell.fill = PatternFill(fill_type=None)

        max_col_letter = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A2:{max_col_letter}{ws.max_row}"

        fixed_widths = {
            "# в33": 40 / 7,
            "Дата рождения": 85 / 7,
            "Место рождения": 285 / 7,
            "Серия-Номер": 95 / 7,
            "Дата выдачи": 100 / 7,
            "Кем выдан": 300 / 7,
            "Код подразделения": 115 / 7,
            "Уникальный идентификатор договора (сделки) БАНКА": 285 / 7,
            "Дата согласия на обработку ПДН (дата договора)": 95 / 7,
            "Статус долга (Операция)": 100 / 7,
            "Дата создания (дата передачи цессии)": 115 / 7,
            "Сумма последнего возрата": 135 / 7,
            "Дата последнего возврата": 125 / 7,
            "ВОЗВРАТ": 7 / 7,
        }

        for col_name, width in fixed_widths.items():
            if col_name in headers:
                ws.column_dimensions[get_column_letter(headers[col_name])].width = width

        auto_headers = [
            "ИД договора",
            "Ключевое поле",
            "Фамилия",
            "Имя",
            "Отчество",
            "Дата согласия на обработку ПДН (дата договора)",
            "Статус долга (Операция)",
            "Дата создания (дата передачи цессии)",
            "Общая сумма долга",
            "Остаток долга",
            "Сумма последнего возрата",
            "Дата последнего возврата",
            "Перевозврат",
            "Цессия",
        ]

        for col_name in auto_headers:
            if col_name in headers and col_name not in fixed_widths:
                col_idx = headers[col_name]
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for row in ws.iter_rows(min_row=2, max_col=col_idx, min_col=col_idx):
                    if row[0].value:
                        max_len = max(max_len, len(str(row[0].value)))
                ws.column_dimensions[col_letter].width = max_len + 2

        wb.save(path)
        logger.success("Кастомизация Excel файла успешно завершена.")
        return True

    except Exception as e:
        logger.exception(f"Ошибка при обработке Excel: {e}")
        return False
