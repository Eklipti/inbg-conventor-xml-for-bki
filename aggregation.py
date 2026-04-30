import sys
from pathlib import Path

import openpyxl
from loguru import logger

from utils import convert_xls_to_xlsx, validate_excel


def process_excel_returns(returns_file_path: Path, main_file_path: Path) -> Path | None:
    """Извлекает данные из файла возвратов, агрегирует их и обновляет основной файл.

    Читает файл возвратов (лист "взносы", любой регистр), ищет заголовки.
    Агрегирует суммы платежей по ключу (Номер ДО, Дата платежа).
    Затем открывает основной файл (лист "Активные"), ищет соответствующие
    записи по "Ключевое поле" (совпадает с Номер ДО). В найденную строку
    записывается дата, общая сумма возврата для группы и пересчитывается
    остаток долга (включая установку статуса "close" и обработку перевозврата).

    Args:
        returns_file_path (Path): Путь к файлу возвратов Excel (файл-донор).
        main_file_path (Path): Путь к основному файлу Excel (файл-реципиент).

    Returns:
        Path | None: Путь к обработанному основному файлу при успехе,
        или None при ошибках валидации, конвертации или отсутствии нужных данных.

    Raises:
        SystemExit: При возникновении критических ошибок в процессе обработки.
    """
    logger.info(f'Слияние данных из "{returns_file_path.name}" в "{main_file_path.name}"')

    ret_path_obj = Path(returns_file_path)
    main_path_obj = Path(main_file_path)

    try:
        if ret_path_obj.suffix.lower() == ".xls":
            logger.info("Обнаружен формат .xls для файла возвратов, запуск конвертации в .xlsx.")
            ret_path_obj = convert_xls_to_xlsx(ret_path_obj)

        logger.trace("Загрузка рабочей книги возвратов.")
        ret_wb = openpyxl.load_workbook(ret_path_obj, data_only=True)

        ret_sheet = None
        for sheet_name in ret_wb.sheetnames:
            if sheet_name.lower() == "взносы":
                ret_sheet = ret_wb[sheet_name]
                break

        if ret_sheet is None:
            logger.error('Лист "взносы" не найден в файле возвратов.')
            return None

        ret_target_headers: list[str] = [
            "Долговое обязательство.Номер ДО",
            "Поступление платежа.Сумма платежа",
            "Поступление платежа.Дата платежа",
        ]
        ret_headers_idx: dict[str, int] = {}

        logger.trace("Поиск индексов заголовков в файле возвратов (строка 1).")
        for r_idx in [1]:
            for row in ret_sheet.iter_rows(min_row=r_idx, max_row=r_idx):
                for cell in row:
                    if cell.value and str(cell.value).strip() in ret_target_headers:
                        ret_headers_idx[str(cell.value).strip()] = cell.column - 1
            if len(ret_headers_idx) == len(ret_target_headers):
                break

        if len(ret_headers_idx) != len(ret_target_headers):
            missing = set(ret_target_headers) - set(ret_headers_idx.keys())
            logger.error(f"В файле возвратов не найдены заголовки: {missing}.")
            return None

        idx_ret_key: int = ret_headers_idx["Долговое обязательство.Номер ДО"]
        idx_ret_sum: int = ret_headers_idx["Поступление платежа.Сумма платежа"]
        idx_ret_date: int = ret_headers_idx["Поступление платежа.Дата платежа"]

        aggregated_sums: dict[tuple[str, str], float] = {}

        logger.info("Начало агрегации данных из файла возвратов.")
        for row_idx, row in enumerate(ret_sheet.iter_rows(min_row=3), start=3):
            cell_key = row[idx_ret_key].value
            cell_sum = row[idx_ret_sum].value
            cell_date = row[idx_ret_date].value

            key_val = str(cell_key).strip() if cell_key is not None else ""
            date_val = str(cell_date).strip() if cell_date is not None else ""
            sum_str = str(cell_sum).strip() if cell_sum is not None else "0"

            if not key_val:
                continue

            try:
                clean_sum = sum_str.replace(" ", "").replace(",", ".")
                sum_float = float(clean_sum)
            except ValueError:
                logger.error(f"Файл возвратов, строка {row_idx}: невозможно преобразовать '{sum_str}' в число.")
                return None

            dict_key = (key_val, date_val)
            if dict_key in aggregated_sums:
                aggregated_sums[dict_key] += sum_float
            else:
                aggregated_sums[dict_key] = sum_float

        logger.debug(f"Агрегация возвратов завершена. Уникальных записей: {len(aggregated_sums)}.")

        if main_path_obj.suffix.lower() == ".xls":
            logger.info("Обнаружен формат .xls для основного файла, запуск конвертации в .xlsx.")
            main_path_obj = convert_xls_to_xlsx(main_path_obj)

        if not validate_excel(main_path_obj):
            logger.warning(f"Основной файл {main_path_obj.name} не прошел валидацию.")
            return None

        logger.debug("Загрузка основного файла Excel.")
        main_wb = openpyxl.load_workbook(main_path_obj)
        main_sheet = main_wb["Активные"]

        main_target_headers: list[str] = [
            "Ключевое поле",
            "Сумма последнего возрата",
            "Дата последнего возврата",
            "Остаток долга",
            "Статус долга (Операция)",
            "Перевозврат",
        ]
        main_headers_idx: dict[str, int] = {}

        logger.trace("Поиск индексов заголовков во второй строке основного файла.")
        for row in main_sheet.iter_rows(min_row=2, max_row=2):
            for cell in row:
                if cell.value and str(cell.value).strip() in main_target_headers:
                    main_headers_idx[str(cell.value).strip()] = cell.column - 1

        if len(main_headers_idx) != len(main_target_headers):
            missing = set(main_target_headers) - set(main_headers_idx.keys())
            logger.error(f"В основном файле не найдены заголовки: {missing}.")
            return None

        idx_main_key: int = main_headers_idx["Ключевое поле"]
        idx_main_sum: int = main_headers_idx["Сумма последнего возрата"]
        idx_main_date: int = main_headers_idx["Дата последнего возврата"]
        idx_main_debt: int = main_headers_idx["Остаток долга"]
        idx_main_status: int = main_headers_idx["Статус долга (Операция)"]
        idx_main_overpayment: int = main_headers_idx["Перевозврат"]

        logger.trace("Создание карты строк основного файла по Ключевому полю.")
        main_key_row_map: dict[str, int] = {}
        for row_idx, row in enumerate(main_sheet.iter_rows(min_row=3), start=3):
            cell_key = row[idx_main_key].value
            key_val = str(cell_key).strip() if cell_key is not None else ""
            if key_val and key_val not in main_key_row_map:
                main_key_row_map[key_val] = row_idx

        logger.info("Начало записи агрегированных сумм в основной файл.")
        for (key, date), total_sum in aggregated_sums.items():
            if key not in main_key_row_map:
                logger.trace(f"Ключ '{key}' не найден в основном файле. Пропуск.")
                continue

            row_idx = main_key_row_map[key]

            formatted_sum: str = f"{total_sum:.2f}".replace(".", ",")
            main_sheet.cell(row=row_idx, column=idx_main_sum + 1, value=formatted_sum)
            main_sheet.cell(row=row_idx, column=idx_main_date + 1, value=date)

            raw_debt = main_sheet.cell(row=row_idx, column=idx_main_debt + 1).value
            debt_str: str = str(raw_debt).strip() if raw_debt is not None else "0"

            try:
                clean_debt: str = debt_str.replace(" ", "").replace(",", ".")
                current_debt: float = float(clean_debt)
            except ValueError:
                logger.error(
                    f"Ошибка в строке {row_idx} основного файла: невозможно преобразовать остаток '{debt_str}'."
                )
                return None

            new_debt: float = current_debt - total_sum

            if new_debt == 0:
                logger.trace(f"Строка {row_idx}: Долг полностью погашен (в ноль).")
                main_sheet.cell(row=row_idx, column=idx_main_status + 1, value="close")
                formatted_new_debt = f"{new_debt:.2f}".replace(".", ",")
                main_sheet.cell(row=row_idx, column=idx_main_debt + 1, value=formatted_new_debt)

            elif new_debt < 0:
                logger.trace(f"Строка {row_idx}: Долг закрыт, есть перевозврат: {new_debt}.")
                main_sheet.cell(row=row_idx, column=idx_main_status + 1, value="close")
                formatted_overpayment: str = f"{new_debt:.2f}".replace(".", ",")
                main_sheet.cell(row=row_idx, column=idx_main_overpayment + 1, value=formatted_overpayment)
                main_sheet.cell(row=row_idx, column=idx_main_debt + 1, value="0,00")

            else:
                logger.trace(f"Строка {row_idx}: Обновлен остаток долга: {new_debt}.")
                formatted_new_debt = f"{new_debt:.2f}".replace(".", ",")
                main_sheet.cell(row=row_idx, column=idx_main_debt + 1, value=formatted_new_debt)

        logger.debug(f"Сохранение изменений в основной файл: {main_path_obj.name}")
        main_wb.save(main_path_obj)
        logger.success("Слияние и нормализация успешно завершены.")

        return main_path_obj

    except Exception as e:
        logger.critical(f"Критический сбой при слиянии файлов: {e}")
        logger.exception("Стек вызовов:")
        sys.exit(1)
